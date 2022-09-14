#Imports
from boto3 import client
from openpyxl import Workbook, load_workbook

#open new workbook
wb = Workbook()

# create_sheet is used to create sheet.
ws2 = wb.create_sheet(title="Title of the sheet")

#filename to be saved as
dest_filename = 'destinationfilename.xlsx'
  
#for the row number
count = 1

s3 = client('s3')

#list_objects_vs help sus overcome the 1000 limit that 'list' command has
paginator = s3.get_paginator('list_objects_v2')


#Input
ws = wb.create_sheet(title="new sheet name")

#Get all the pages from the bucket
pages = paginator.paginate(Bucket='AWS-S3-Bucketname', Prefix='folderName/')

#Add the column to the sheet
ws.append(["Key", "ETag", "Size", "StorageClass"])

#For each page
for page in pages:

    #For each item in the page
    for obj in page['Contents']:

        #Extracting the items from the filepath "folderName/"
        Key = obj['Key'].split('/')
        item_ = Key[-1].split('-')

        #print once every 1000 records to check on the progress
        if count % 1000 == 0:
            print(count, [item_[0], obj["ETag"], obj["Size"], obj["StorageClass"]])
        
        #appending the new row for each SO
        ws.append([item_[0], obj["ETag"], obj["Size"], obj["StorageClass"]])

        #Incrementing the count for next row.
        count += 1

#Printing total count after the run is complete
print(count)

#Save and exit the Excel file with the given filename
wb.save(filename = dest_filename)